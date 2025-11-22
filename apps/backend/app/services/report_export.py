from __future__ import annotations

from io import BytesIO
from enum import Enum

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import LETTER

    # from reportlab.pdfgen import canvas  # unused
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    REPORTLAB_AVAILABLE = True
except Exception:  # pragma: no cover - during minimal envs
    REPORTLAB_AVAILABLE = False


class ReportMode(str, Enum):
    """Export mode for reports."""

    full = "full"
    summary = "summary"
    unknowns = "unknowns"


# --- Excel sheet builders using openpyxl --------------------------------------


def add_summary_sheet(
    wb: "Workbook", month: str, summary: dict, unknown_count: int, unknown_amount: float
) -> None:
    """Add Summary sheet with overview metrics."""
    ws = wb.create_sheet("Summary")

    # Title
    ws["A1"] = "LedgerMind — Monthly Summary"
    ws["A1"].font = Font(bold=True, size=14)

    # Month info
    ws["A2"] = "Month"
    ws["B2"] = month

    # Table header (row 4)
    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    # Metrics
    ws.append(["Total income", round(float(summary.get("total_income", 0.0)), 2)])
    ws.append(["Total spend", round(float(summary.get("total_spend", 0.0)), 2)])
    ws.append(["Net", round(float(summary.get("net", 0.0)), 2)])
    ws.append(["Unknown spend", round(float(unknown_amount), 2)])
    ws.append(["Unknown txns", unknown_count])

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def add_categories_sheet(wb: "Workbook", categories: list[dict]) -> None:
    """Add Categories sheet with category breakdown."""
    ws = wb.create_sheet("Categories")

    # Header
    headers = [
        "category_slug",
        "category_label",
        "txn_count",
        "total_amount",
        "pct_of_spend",
    ]
    ws.append(headers)

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for cat in categories:
        ws.append(
            [
                cat.get("slug", cat.get("category", "")),
                cat.get("label", cat.get("name", cat.get("category", ""))),
                cat.get("txn_count", cat.get("count", 0)),
                round(float(cat.get("amount", cat.get("spend", 0.0))), 2),
                round(float(cat.get("pct_of_spend", 0.0)), 2),
            ]
        )

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12


def add_merchants_sheet(wb: "Workbook", merchants: list[dict]) -> None:
    """Add Merchants sheet with merchant breakdown (using canonical grouping)."""
    ws = wb.create_sheet("Merchants")

    # Header
    headers = [
        "merchant_display",
        "merchant_canonical",
        "txn_count",
        "total_amount",
        "top_category_slug",
        "top_category_label",
    ]
    ws.append(headers)

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for m in merchants:
        ws.append(
            [
                m.get("merchant_display", m.get("merchant", "")),
                m.get("merchant_canonical", m.get("canonical", "")),
                m.get("n", m.get("txn_count", 0)),
                round(float(m.get("amount", m.get("total_amount", 0.0))), 2),
                m.get("top_category_slug", m.get("category", "")),
                m.get("top_category_label", m.get("category_label", "")),
            ]
        )

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def add_transactions_sheet(wb: "Workbook", transactions: list[dict]) -> None:
    """Add Transactions sheet with all transaction details."""
    ws = wb.create_sheet("Transactions")

    # Header
    headers = [
        "date",
        "merchant",
        "description",
        "amount",
        "category_label",
        "category_slug",
    ]
    ws.append(headers)

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for txn in transactions:
        ws.append(
            [
                txn.get("date", ""),
                txn.get("merchant", ""),
                txn.get("description", ""),
                round(float(txn.get("amount", 0.0)), 2),
                txn.get("category_label", txn.get("category", "")),
                txn.get("category_slug", txn.get("category", "")),
            ]
        )

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20


def add_unknowns_sheet(
    wb: "Workbook", month: str, unknown_transactions: list[dict]
) -> None:
    """Add Unknowns sheet with only uncategorized transactions."""
    ws = wb.create_sheet("Unknowns")

    # Title
    ws["A1"] = f"Uncategorized Transactions — {month}"
    ws["A1"].font = Font(bold=True, size=14)

    # Header (row 3)
    ws["A3"] = "date"
    ws["B3"] = "merchant"
    ws["C3"] = "description"
    ws["D3"] = "amount"
    ws["E3"] = "category_label"
    ws["F3"] = "category_slug"

    # Make header bold
    for cell in ws[3]:
        cell.font = Font(bold=True)

    # Data rows
    for txn in unknown_transactions:
        ws.append(
            [
                txn.get("date", ""),
                txn.get("merchant", ""),
                txn.get("description", ""),
                round(float(txn.get("amount", 0.0)), 2),
                txn.get("category_label", txn.get("category", "")),
                txn.get("category_slug", txn.get("category", "")),
            ]
        )

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20


def build_excel_bytes(
    summary: dict,
    merchants: list[dict],
    categories: list[dict],
    flows: dict | None = None,
    trends: dict | list | None = None,
    txns_df=None,  # Legacy pandas support (deprecated)
    split_txns_alpha: bool = False,
    # New structured parameters
    mode: ReportMode = ReportMode.full,
    transactions: list[dict] | None = None,
    unknown_transactions: list[dict] | None = None,
) -> bytes:
    """Build an Excel workbook with structured sheets based on mode.

    Args:
        summary: Month summary dict with total_income, total_spend, net
        merchants: List of merchant dicts with canonical grouping
        categories: List of category dicts
        flows: Optional flows dict (legacy)
        trends: Optional trends data (legacy)
        txns_df: Optional pandas DataFrame (deprecated, use transactions list)
        split_txns_alpha: Whether to split transactions A-M / N-Z (legacy)
        mode: Export mode (full/summary/unknowns)
        transactions: List of transaction dicts for full mode
        unknown_transactions: List of unknown transaction dicts for unknowns mode

    Returns:
        Excel workbook as bytes
    """
    if not OPENPYXL_AVAILABLE:
        # Fallback to pandas-based builder for backward compatibility
        return _build_excel_bytes_pandas(
            summary, merchants, categories, flows, trends, txns_df, split_txns_alpha
        )

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

    # Calculate unknown stats
    unknown_count = len(unknown_transactions) if unknown_transactions else 0
    unknown_amount = (
        sum(float(t.get("amount", 0)) for t in unknown_transactions)
        if unknown_transactions
        else 0.0
    )

    month_str = summary.get("month", "")

    # Build sheets based on mode
    if mode == ReportMode.full:
        add_summary_sheet(wb, month_str, summary, unknown_count, unknown_amount)
        add_categories_sheet(wb, categories)
        add_merchants_sheet(wb, merchants)
        if transactions:
            add_transactions_sheet(wb, transactions)
    elif mode == ReportMode.summary:
        add_summary_sheet(wb, month_str, summary, unknown_count, unknown_amount)
    elif mode == ReportMode.unknowns:
        if unknown_transactions:
            add_unknowns_sheet(wb, month_str, unknown_transactions)
        else:
            # Fallback: empty unknowns sheet if no data
            add_unknowns_sheet(wb, month_str, [])
    else:
        # Defensive fallback: summary mode
        add_summary_sheet(wb, month_str, summary, unknown_count, unknown_amount)

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_excel_bytes_pandas(
    summary: dict,
    merchants: list[dict],
    categories: list[dict],
    flows: dict | None = None,
    trends: dict | list | None = None,
    txns_df=None,
    split_txns_alpha: bool = False,
) -> bytes:
    """Legacy pandas-based Excel builder (fallback when openpyxl unavailable)."""
    import pandas as pd

    df_summary = pd.DataFrame(
        [
            {
                "Month": summary.get("month"),
                "Total Spend": round(float(summary.get("total_spend", 0.0)), 2),
                "Total Income": round(float(summary.get("total_income", 0.0)), 2),
                "Net": round(float(summary.get("net", 0.0)), 2),
            }
        ]
    )
    df_categories = pd.DataFrame(categories or [])
    df_merchants = pd.DataFrame(merchants or [])

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        if not df_categories.empty:
            df_categories.to_excel(writer, index=False, sheet_name="Categories")
        if not df_merchants.empty:
            df_merchants.to_excel(writer, index=False, sheet_name="TopMerchants")
        if flows:
            pd.DataFrame([flows]).to_excel(writer, index=False, sheet_name="Flows")
        if trends:
            tdf = pd.DataFrame(
                trends if isinstance(trends, list) else trends.get("trends", [])
            )
            if not tdf.empty:
                tdf.to_excel(writer, index=False, sheet_name="Trends")
        if txns_df is not None:
            try:
                if not txns_df.empty:
                    if split_txns_alpha and "merchant" in txns_df.columns:
                        df = txns_df.copy()
                        first = df["merchant"].fillna("").str.strip().str.upper().str[0]
                        am = df[first.isin(list("ABCDEFGHIJKLM"))]
                        nz = df[
                            first.isin(list("NOPQRSTUVWXYZ"))
                            | (~first.str.match(r"[A-Z]"))
                        ]
                        if not am.empty:
                            am.to_excel(
                                writer, index=False, sheet_name="Transactions A-M"
                            )
                        if not nz.empty:
                            nz.to_excel(
                                writer, index=False, sheet_name="Transactions N-Z"
                            )
                    else:
                        txns_df.to_excel(writer, index=False, sheet_name="Transactions")
            except Exception:
                pass
        for sheet in ("Summary", "Categories", "TopMerchants", "Flows", "Trends"):
            if sheet in writer.sheets:
                ws = writer.sheets[sheet]
                ws.set_column(0, 20, 18, cell_format=None)
    return bio.getvalue()


def build_pdf_bytes(
    summary: dict,
    merchants: list[dict],
    categories: list[dict] | None = None,
    flows=None,
    trends=None,
) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab is not installed in this environment")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LETTER, title="Monthly Finance Report")
    styles = getSampleStyleSheet()
    story = []

    title = f"Monthly Finance Report — {summary.get('month') or (summary.get('start') or '')}→{summary.get('end') or ''}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 12))

    # Summary stats
    story.append(
        Paragraph(
            f"Total Spend: ${summary.get('total_spend', 0):.2f}", styles["Normal"]
        )
    )
    story.append(
        Paragraph(
            f"Total Income: ${summary.get('total_income', 0):.2f}", styles["Normal"]
        )
    )
    story.append(Paragraph(f"Net: ${summary.get('net', 0):.2f}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Merchants table
    if merchants:
        story.append(Paragraph("Top Merchants (by spend)", styles["Heading2"]))
        m_rows = [["Merchant", "Spend", "Txns"]] + [
            [m.get("merchant"), f"${m.get('amount', 0):.2f}", str(m.get("n", 0))]
            for m in merchants[:15]
        ]
        mt = Table(m_rows, hAlign="LEFT", colWidths=[280, 90, 60])
        mt.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(mt)
        story.append(Spacer(1, 12))

    # Categories table
    if categories:
        story.append(Paragraph("Top Categories (by spend)", styles["Heading2"]))
        c_rows = [["Category", "Spend"]] + [
            [
                c.get("category") or c.get("name"),
                f"${c.get('spend', c.get('amount', 0)):.2f}",
            ]
            for c in categories[:15]
        ]
        ct = Table(c_rows, hAlign="LEFT", colWidths=[280, 90])
        ct.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(ct)
        story.append(Spacer(1, 12))

    doc.build(story)
    return buf.getvalue()
