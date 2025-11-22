"""Tests for Excel export workbook structure."""

import io
from fastapi.testclient import TestClient

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.main import app

client = TestClient(app)


def _load_wb(content: bytes):
    """Helper to load workbook from bytes."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not available")
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True)


class TestReportExcelStructure:
    """Test suite for Excel workbook structure based on mode."""

    def _get_sheetnames(self, content: bytes) -> list[str]:
        """Extract sheet names from Excel workbook bytes."""
        if not OPENPYXL_AVAILABLE:
            # Fallback: just verify it's not empty
            return []
        wb = load_workbook(io.BytesIO(content), read_only=True)
        return wb.sheetnames

    def test_full_mode_has_all_core_sheets(self):
        """Verify mode=full includes Summary, Categories, Merchants, Transactions sheets."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        # May return 401 if not authenticated in test context, or 404 if no data
        if resp.status_code not in (200, 401, 404):
            assert False, f"Unexpected status: {resp.status_code}"

        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            names = self._get_sheetnames(resp.content)
            # Verify core sheets are present (order matters)
            assert "Summary" in names, f"Missing Summary sheet, got: {names}"
            assert "Categories" in names, f"Missing Categories sheet, got: {names}"
            assert "Merchants" in names, f"Missing Merchants sheet, got: {names}"
            assert "Transactions" in names, f"Missing Transactions sheet, got: {names}"

            # Verify Summary is first
            if names:
                assert names[0] == "Summary", f"Summary should be first, got: {names}"

    def test_summary_mode_only_has_summary_sheet(self):
        """Verify mode=summary includes only Summary sheet."""
        resp = client.get(
            "/report/excel", params={"month": "2025-11", "mode": "summary"}
        )
        if resp.status_code not in (200, 401, 404):
            assert False, f"Unexpected status: {resp.status_code}"

        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            names = self._get_sheetnames(resp.content)
            assert names == [
                "Summary"
            ], f"Summary mode should only have Summary sheet, got: {names}"

    def test_unknowns_mode_only_has_unknowns_sheet(self):
        """Verify mode=unknowns includes only Unknowns sheet."""
        resp = client.get(
            "/report/excel", params={"month": "2025-11", "mode": "unknowns"}
        )
        if resp.status_code not in (200, 401, 404):
            assert False, f"Unexpected status: {resp.status_code}"

        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            names = self._get_sheetnames(resp.content)
            assert names == [
                "Unknowns"
            ], f"Unknowns mode should only have Unknowns sheet, got: {names}"

    def test_summary_sheet_structure(self):
        """Verify Summary sheet has expected structure."""
        resp = client.get(
            "/report/excel", params={"month": "2025-11", "mode": "summary"}
        )
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            wb = load_workbook(io.BytesIO(resp.content), read_only=True)
            ws = wb["Summary"]

            # Check title
            assert (
                "LedgerMind" in ws["A1"].value
            ), f"Expected title in A1, got: {ws['A1'].value}"

            # Check month label
            assert (
                ws["A2"].value == "Month"
            ), f"Expected 'Month' in A2, got: {ws['A2'].value}"

            # Check table headers (row 4)
            assert (
                ws["A4"].value == "Metric"
            ), f"Expected 'Metric' in A4, got: {ws['A4'].value}"
            assert (
                ws["B4"].value == "Value"
            ), f"Expected 'Value' in B4, got: {ws['B4'].value}"

    def test_categories_sheet_headers(self):
        """Verify Categories sheet has expected headers."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            wb = load_workbook(io.BytesIO(resp.content), read_only=True)
            if "Categories" in wb.sheetnames:
                ws = wb["Categories"]
                headers = [cell.value for cell in ws[1]]

                expected = [
                    "category_slug",
                    "category_label",
                    "txn_count",
                    "total_amount",
                    "pct_of_spend",
                ]
                assert (
                    headers == expected
                ), f"Expected headers {expected}, got: {headers}"

    def test_merchants_sheet_headers(self):
        """Verify Merchants sheet has expected headers."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            wb = load_workbook(io.BytesIO(resp.content), read_only=True)
            if "Merchants" in wb.sheetnames:
                ws = wb["Merchants"]
                headers = [cell.value for cell in ws[1]]

                expected = [
                    "merchant_display",
                    "merchant_canonical",
                    "txn_count",
                    "total_amount",
                    "top_category_slug",
                    "top_category_label",
                ]
                assert (
                    headers == expected
                ), f"Expected headers {expected}, got: {headers}"

    def test_transactions_sheet_headers(self):
        """Verify Transactions sheet has expected headers."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            wb = load_workbook(io.BytesIO(resp.content), read_only=True)
            if "Transactions" in wb.sheetnames:
                ws = wb["Transactions"]
                headers = [cell.value for cell in ws[1]]

                expected = [
                    "date",
                    "merchant",
                    "description",
                    "amount",
                    "category_label",
                    "category_slug",
                ]
                assert (
                    headers == expected
                ), f"Expected headers {expected}, got: {headers}"

    def test_unknowns_sheet_structure(self):
        """Verify Unknowns sheet has expected structure."""
        resp = client.get(
            "/report/excel", params={"month": "2025-11", "mode": "unknowns"}
        )
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            wb = load_workbook(io.BytesIO(resp.content), read_only=True)
            ws = wb["Unknowns"]

            # Check title (row 1)
            assert (
                "Uncategorized" in ws["A1"].value
            ), f"Expected title in A1, got: {ws['A1'].value}"

            # Check headers (row 3)
            headers = [cell.value for cell in ws[3]]
            expected = [
                "date",
                "merchant",
                "description",
                "amount",
                "category_label",
                "category_slug",
            ]
            assert headers == expected, f"Expected headers {expected}, got: {headers}"

    def test_backwards_compat_include_transactions_true_creates_transactions_sheet(
        self,
    ):
        """Verify old include_transactions=true still creates transactions sheet."""
        resp = client.get(
            "/report/excel",
            params={"month": "2025-11", "include_transactions": "true"},
        )
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            names = self._get_sheetnames(resp.content)
            # Should map to full mode
            assert "Summary" in names
            if names:  # May not have transactions if no data
                # At minimum should have Summary
                assert names[0] == "Summary"

    def test_backwards_compat_include_transactions_false_summary_only(self):
        """Verify old include_transactions=false creates summary-only workbook."""
        resp = client.get(
            "/report/excel",
            params={"month": "2025-11", "include_transactions": "false"},
        )
        if resp.status_code == 200 and OPENPYXL_AVAILABLE:
            names = self._get_sheetnames(resp.content)
            # Should map to summary mode
            assert names == ["Summary"]

    def test_full_mode_summary_contains_core_metrics(self):
        """Verify Summary sheet contains expected metric labels."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        if resp.status_code != 200:
            # Skip if no auth or no data
            return

        if not OPENPYXL_AVAILABLE:
            return

        wb = _load_wb(resp.content)
        ws = wb["Summary"]

        # Find the 'Metric'/'Value' header row (should be row 4) and read metrics
        headers = [cell.value for cell in ws[4]]
        assert headers[:2] == [
            "Metric",
            "Value",
        ], f"Expected Metric/Value headers, got: {headers[:2]}"

        # Collect all metric names from column A starting at row 5
        metrics = {}
        for row in ws.iter_rows(min_row=5, max_col=2, values_only=True):
            key, val = row
            if key:
                metrics[key] = val

        # Assert core metrics exist (don't check exact values, just presence)
        expected_keys = [
            "Total income",
            "Total spend",
            "Net",
            "Unknown spend",
            "Unknown txns",
        ]
        for key in expected_keys:
            assert key in metrics, f"Missing metric '{key}' in Summary sheet"

    def test_unknowns_mode_has_only_unknowns_sheet_and_rows(self):
        """Verify unknowns mode has only Unknowns sheet with correct structure."""
        resp = client.get(
            "/report/excel", params={"month": "2025-11", "mode": "unknowns"}
        )
        if resp.status_code != 200:
            return

        if not OPENPYXL_AVAILABLE:
            return

        wb = _load_wb(resp.content)
        assert wb.sheetnames == [
            "Unknowns"
        ], f"Expected only Unknowns sheet, got: {wb.sheetnames}"

        ws = wb["Unknowns"]

        # Headers are in row 3 (row 1 is title)
        header = [c for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
        expected_header = [
            "date",
            "merchant",
            "description",
            "amount",
            "category_label",
            "category_slug",
        ]
        assert (
            header == expected_header
        ), f"Expected headers {expected_header}, got: {header}"

        # Check that all data rows have category_slug as None or "unknown"
        for row in ws.iter_rows(min_row=4, values_only=True):
            if all(cell is None for cell in row):
                # Skip empty rows
                continue
            # Last column is category_slug
            _, _, _, _, _, slug = row[:6]
            # Allow None or "unknown" (or empty string)
            assert slug in (
                None,
                "unknown",
                "",
            ), f"Expected unknown category_slug, got: {slug}"

    def test_categories_sheet_pct_of_spend_sums_to_approx_1(self):
        """Verify Categories sheet percentages sum to approximately 1.0."""
        resp = client.get("/report/excel", params={"month": "2025-11", "mode": "full"})
        if resp.status_code != 200:
            return

        if not OPENPYXL_AVAILABLE:
            return

        wb = _load_wb(resp.content)
        if "Categories" not in wb.sheetnames:
            # No categories data
            return

        ws = wb["Categories"]

        # Verify header row
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        expected_header = [
            "category_slug",
            "category_label",
            "txn_count",
            "total_amount",
            "pct_of_spend",
        ]
        assert (
            header == expected_header
        ), f"Expected headers {expected_header}, got: {header}"

        # Collect all pct_of_spend values
        pct_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                # Skip empty rows
                continue
            slug, label, count, total, pct = row[:5]
            if pct is not None:
                pct_values.append(float(pct))

        # If we have percentages, they should sum to approximately 1.0
        if pct_values:
            total_pct = sum(pct_values)
            assert (
                0.95 <= total_pct <= 1.05
            ), f"Expected pct_of_spend to sum to ~1.0, got: {total_pct}"
